#!/usr/bin/env python3
"""
Proofreads the text of a .xlsx (text cells only, ignoring numbers and formulas) and FLAGS
likely mistakes so the model sees them and fixes them before delivering. Extracts with openpyxl
and applies light heuristics: encoding artifacts (mojibake), doubled words, and
space-before-punctuation. Own implementation (MIT).

Usage:
    python -X utf8 spellcheck.py workbook.xlsx

Always run with `-X utf8` on Windows. This is an ASSISTANT, not an infallible checker.
"""
import re
import sys

from openpyxl import load_workbook

MOJIBAKE = ["Ã", "Â", "Ð", "�", "â€", "Ã©", "Ã±", "Ã¡", "Ã³", "Ãº"]
DOUBLED = re.compile(r"\b(\w+)\s+\1\b", re.IGNORECASE)
SPACE_BEFORE_PUNCT = re.compile(r"\s+[,.;:!?]")


def check_cell(location: str, text: str):
    findings = []
    for mark in MOJIBAKE:
        if mark in text:
            findings.append((location, f"mojibake '{mark}': {text[:60]}"))
            break
    for m in DOUBLED.finditer(text):
        findings.append((location, f"doubled word '{m.group(0)}': {text[:60]}"))
    if SPACE_BEFORE_PUNCT.search(text):
        findings.append((location, f"space before punctuation: {text[:60]}"))
    return findings


def main():
    if len(sys.argv) != 2:
        print("usage: python -X utf8 spellcheck.py <workbook.xlsx>")
        sys.exit(1)

    wb = load_workbook(sys.argv[1], data_only=False, read_only=True)
    findings, text_cells = [], []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and not v.startswith("="):
                    loc = f"{name}!{cell.coordinate}"
                    text_cells.append((loc, v))
                    findings.extend(check_cell(loc, v))
    wb.close()

    seen, unique = set(), []
    for f in findings:
        if f not in seen:
            seen.add(f); unique.append(f)

    print("=" * 70)
    print(f"PROOFREAD — {len(unique)} possible issue(s) flagged")
    print("=" * 70)
    if unique:
        for loc, msg in unique:
            print(f"  {loc}: {msg}")
    else:
        print("  (no automatic suspects — still READ the text cells below)")

    print("\n" + "-" * 70)
    print("TEXT CELLS (read them for mistakes the heuristics can't catch):")
    print("-" * 70)
    for loc, v in text_cells:
        print(f"  {loc}: {v}")


if __name__ == "__main__":
    main()
