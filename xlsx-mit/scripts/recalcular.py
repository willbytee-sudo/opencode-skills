#!/usr/bin/env python3
"""
Recalcula las fórmulas de un .xlsx con LibreOffice (headless) y escanea todas las celdas
en busca de errores de Excel (#REF!, #DIV/0!, ...). Devuelve JSON con el detalle.
Implementación propia (MIT). Requiere openpyxl (✅) y LibreOffice (soffice).

Uso:
    python -X utf8 recalcular.py libro.xlsx [timeout_segundos]

Correr SIEMPRE con `-X utf8` en Windows.

Si NO tienes LibreOffice: openpyxl escribe las fórmulas pero no sus valores; Excel las
calcula al abrir el archivo. recalcular.py solo hace falta para obtener/validar esos valores
aquí mismo. Instala LibreOffice y agrega soffice al PATH (o exporta SOFFICE_BIN).
"""
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ERRORES = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Recalc" script:language="StarBasic">
Sub RecalcularYGuardar()
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>"""


def buscar_soffice():
    c = os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or shutil.which("soffice.exe")
    if c:
        return c
    for p in (r"C:\Program Files\LibreOffice\program\soffice.exe",
              r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"):
        if Path(p).exists():
            return p
    return None


def instalar_macro(soffice: str) -> bool:
    appdata = os.environ.get("APPDATA", str(Path.home() / "AppData" / "Roaming"))
    carpeta = Path(appdata) / "LibreOffice" / "4" / "user" / "basic" / "Standard"
    archivo = carpeta / "Recalc.xba"
    if archivo.exists() and "RecalcularYGuardar" in archivo.read_text(encoding="utf-8", errors="ignore"):
        return True
    if not carpeta.exists():
        try:
            subprocess.run([soffice, "--headless", "--terminate_after_init"],
                           capture_output=True, timeout=30)
        except Exception:  # noqa: BLE001
            pass
        carpeta.mkdir(parents=True, exist_ok=True)
    try:
        archivo.write_text(MACRO, encoding="utf-8")
        # registrar el módulo en script.xlb si no está (LibreOffice lo suele autodetectar)
        return True
    except Exception:  # noqa: BLE001
        return False


def escanear_errores(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    detalle = {e: [] for e in ERRORES}
    total = 0
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for e in ERRORES:
                        if e in cell.value:
                            detalle[e].append(f"{name}!{cell.coordinate}"); total += 1
                            break
    wb.close()
    wbf = load_workbook(path, data_only=False)
    formulas = sum(1 for n in wbf.sheetnames for row in wbf[n].iter_rows()
                   for c in row if isinstance(c.value, str) and c.value.startswith("="))
    wbf.close()
    res = {"status": "success" if total == 0 else "errors_found",
           "total_errors": total, "total_formulas": formulas, "error_summary": {}}
    for e, locs in detalle.items():
        if locs:
            res["error_summary"][e] = {"count": len(locs), "locations": locs[:20]}
    return res


def recalcular(path, timeout=30) -> dict:
    if not Path(path).exists():
        return {"error": f"El archivo {path} no existe"}
    soffice = buscar_soffice()
    if not soffice:
        return {"error": "LibreOffice (soffice) no está instalado. Instálalo y agrégalo al "
                         "PATH, o exporta SOFFICE_BIN. (Excel calcula las fórmulas al abrir el archivo.)"}
    if not instalar_macro(soffice):
        return {"error": "No se pudo instalar la macro de LibreOffice"}
    cmd = [soffice, "--headless", "--norestore",
           "vnd.sun.star.script:Standard.Recalc.RecalcularYGuardar?language=Basic&location=application",
           str(Path(path).absolute())]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice excedió el timeout de {timeout}s"}
    if r.returncode != 0:
        return {"error": r.stderr or "Error desconocido durante el recálculo"}
    try:
        return escanear_errores(path)
    except Exception as e:  # noqa: BLE001
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("uso: python -X utf8 recalcular.py <libro.xlsx> [timeout_segundos]")
        sys.exit(1)
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalcular(sys.argv[1], timeout), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
