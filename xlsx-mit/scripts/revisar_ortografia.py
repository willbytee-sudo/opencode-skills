#!/usr/bin/env python3
"""
Revisa la ortografía del texto de un .xlsx (solo celdas de texto, ignora números y
fórmulas) y SEÑALA posibles faltas para que el modelo las vea y las corrija antes de
entregar. Extrae con openpyxl y aplica heurísticas para el español: mojibake, tildes
perdidas, terminaciones -ción/-sión/-tión sin tilde y palabras que sin tilde no existen.
Implementación propia (MIT).

Uso:
    python -X utf8 revisar_ortografia.py libro.xlsx

Correr SIEMPRE con `-X utf8` en Windows. Es un ASISTENTE, no un corrector infalible.
"""
import re
import sys

from openpyxl import load_workbook

SIEMPRE_TILDE = {
    "tambien": "también", "asi": "así", "aqui": "aquí", "alli": "allí", "aca": "acá",
    "alla": "allá", "segun": "según", "despues": "después", "ademas": "además",
    "quizas": "quizás", "atras": "atrás", "detras": "detrás", "jamas": "jamás",
    "numero": "número", "numeros": "números", "codigo": "código", "pagina": "página",
    "paginas": "páginas", "linea": "línea", "lineas": "líneas", "titulo": "título",
    "parrafo": "párrafo", "capitulo": "capítulo", "indice": "índice", "metodo": "método",
    "analisis": "análisis", "sintesis": "síntesis", "rapido": "rápido", "practico": "práctico",
    "basico": "básico", "publico": "público", "politica": "política", "economia": "economía",
    "energia": "energía", "categoria": "categoría", "dia": "día", "dias": "días",
    "ultimo": "último", "proximo": "próximo", "minimo": "mínimo", "maximo": "máximo",
    "facil": "fácil", "dificil": "difícil", "util": "útil", "debil": "débil",
    "arbol": "árbol", "azucar": "azúcar", "telefono": "teléfono", "sabado": "sábado",
    "miercoles": "miércoles", "album": "álbum",
}
MOJIBAKE = ["Ã", "Â", "Ð", "�", "â€", "Ã³", "Ã±", "Ã©", "Ã¡", "Ãº", "Ã­"]
VOCAL_MOJI = re.compile(r"[a-záéíóúñ][AEIOU][a-záéíóúñ]")
SUFIJO_ION = re.compile(r"\b[a-zñ]*[bcdfghjklmnpqrstvwxyz]ion\b", re.IGNORECASE)
TOKEN = re.compile(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+")


def revisar_celda(ubicacion: str, texto: str):
    hallazgos = []
    low = texto.lower()
    for marca in MOJIBAKE:
        if marca in texto:
            hallazgos.append((ubicacion, f"mojibake '{marca}': {texto[:60]}"))
            break
    for m in VOCAL_MOJI.finditer(texto):
        hallazgos.append((ubicacion, f"posible tilde perdida cerca de '{m.group(0)}': {texto[:60]}"))
    for w in TOKEN.findall(low):
        if w in SIEMPRE_TILDE:
            hallazgos.append((ubicacion, f"'{w}' -> '{SIEMPRE_TILDE[w]}'"))
    for m in SUFIJO_ION.finditer(low):
        w = m.group(0)
        hallazgos.append((ubicacion, f"'{w}' -> '{w[:-3]}ión' (¿falta tilde?)"))
    return hallazgos


def main():
    if len(sys.argv) != 2:
        print("uso: python -X utf8 revisar_ortografia.py <libro.xlsx>")
        sys.exit(1)

    wb = load_workbook(sys.argv[1], data_only=False, read_only=True)
    hallazgos = []
    celdas_texto = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                v = cell.value
                # solo texto que NO sea fórmula
                if isinstance(v, str) and not v.startswith("="):
                    ubic = f"{name}!{cell.coordinate}"
                    celdas_texto.append((ubic, v))
                    hallazgos.extend(revisar_celda(ubic, v))
    wb.close()

    # de-duplicar
    vistos, unicos = set(), []
    for h in hallazgos:
        if h not in vistos:
            vistos.add(h); unicos.append(h)

    print("=" * 70)
    print(f"REVISIÓN ORTOGRÁFICA — {len(unicos)} posible(s) falta(s) señaladas")
    print("=" * 70)
    if unicos:
        for ubic, msg in unicos:
            print(f"  {ubic}: {msg}")
    else:
        print("  (sin sospechosos automáticos — igual REVISA los textos de abajo)")

    print("\n" + "-" * 70)
    print("CELDAS DE TEXTO (léelas buscando faltas que la heurística no detecta):")
    print("-" * 70)
    for ubic, v in celdas_texto:
        print(f"  {ubic}: {v}")


if __name__ == "__main__":
    main()
