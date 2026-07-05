#!/usr/bin/env python3
"""
Extrae el contenido de un .xlsx/.xlsm como filas separadas por tabulador, una sección
`## Hoja:` por hoja. Usa openpyxl. Reemplaza al binario `extract-text` del sandbox de
Anthropic. Implementación propia (MIT).

Uso:
    python -X utf8 extraer_texto.py libro.xlsx [--valores]

Por defecto muestra las FÓRMULAS tal cual. Con `--valores` muestra los valores cacheados
(data_only), que solo existen si el archivo fue calculado por Excel/LibreOffice (ver recalcular.py).

Correr SIEMPRE con `-X utf8` en Windows.
"""
import sys

from openpyxl import load_workbook


def main():
    args = sys.argv[1:]
    valores = "--valores" in args
    args = [a for a in args if a != "--valores"]
    if len(args) != 1:
        print("uso: python -X utf8 extraer_texto.py <libro.xlsx> [--valores]")
        sys.exit(1)

    wb = load_workbook(args[0], data_only=valores, read_only=True)
    for name in wb.sheetnames:
        print(f"## Hoja: {name}")
        for row in wb[name].iter_rows(values_only=True):
            cells = list(row)
            while cells and (cells[-1] is None or cells[-1] == ""):
                cells.pop()
            if cells:
                print("\t".join("" if c is None else str(c) for c in cells))
        print()
    wb.close()


if __name__ == "__main__":
    main()
