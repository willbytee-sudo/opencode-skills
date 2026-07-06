#!/usr/bin/env python3
"""
Extracts the content of a .xlsx/.xlsm as tab-separated rows, one `## Sheet:` section per sheet.
Uses openpyxl. Replaces the sandbox-only `extract-text` binary. Own implementation (MIT).

Usage:
    python -X utf8 extract_text.py workbook.xlsx [--values]

By default it shows the FORMULAS as-is. With `--values` it shows the cached values (data_only),
which only exist if the file was calculated by Excel/LibreOffice (see recalc.py).

Always run with `-X utf8` on Windows.
"""
import sys

from openpyxl import load_workbook


def main():
    args = sys.argv[1:]
    values = "--values" in args
    args = [a for a in args if a != "--values"]
    if len(args) != 1:
        print("usage: python -X utf8 extract_text.py <workbook.xlsx> [--values]")
        sys.exit(1)

    wb = load_workbook(args[0], data_only=values, read_only=True)
    for name in wb.sheetnames:
        print(f"## Sheet: {name}")
        for row in wb[name].iter_rows(values_only=True):
            cells = list(row)
            while cells and (cells[-1] is None or cells[-1] == ""):
                cells.pop()
            if cells:
                print("\t".join("" if c is None else str(c) for c in cells))
        print()
    wb.close()


if __name__ == "__main__":
    main()
