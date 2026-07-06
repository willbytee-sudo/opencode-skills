#!/usr/bin/env python3
"""
Recalculates the formulas of a .xlsx with LibreOffice (headless) and scans all cells for Excel
errors (#REF!, #DIV/0!, ...). Returns JSON with the details. Own implementation (MIT). Requires
openpyxl (✅) and LibreOffice (soffice).

Usage:
    python -X utf8 recalc.py workbook.xlsx [timeout_seconds]

Always run with `-X utf8` on Windows.

If you do NOT have LibreOffice: openpyxl writes the formulas but not their values; Excel
calculates them when you open the file. recalc.py is only needed to obtain/validate those values
here. Install LibreOffice and add soffice to PATH (or set SOFFICE_BIN).
"""
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Recalc" script:language="StarBasic">
Sub RecalcAndSave
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>"""


def find_soffice():
    c = os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or shutil.which("soffice.exe")
    if c:
        return c
    for p in (r"C:\Program Files\LibreOffice\program\soffice.exe",
              r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"):
        if Path(p).exists():
            return p
    return None


def macro_dir_windows() -> Path:
    appdata = os.environ.get("APPDATA", str(Path.home() / "AppData" / "Roaming"))
    return Path(appdata) / "LibreOffice" / "4" / "user" / "basic" / "Standard"


def setup_macro(soffice: str) -> bool:
    d = macro_dir_windows()
    f = d / "Recalc.xba"
    if f.exists() and "RecalcAndSave" in f.read_text(encoding="utf-8", errors="ignore"):
        return True
    if not d.exists():
        try:
            subprocess.run([soffice, "--headless", "--terminate_after_init"], capture_output=True, timeout=30)
        except Exception:  # noqa: BLE001
            pass
        d.mkdir(parents=True, exist_ok=True)
    try:
        f.write_text(MACRO, encoding="utf-8")
        return True
    except Exception:  # noqa: BLE001
        return False


def scan_errors(filename: str) -> dict:
    wb = load_workbook(filename, data_only=True)
    details = {e: [] for e in ERRORS}
    total = 0
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for e in ERRORS:
                        if e in cell.value:
                            details[e].append(f"{name}!{cell.coordinate}"); total += 1
                            break
    wb.close()
    wbf = load_workbook(filename, data_only=False)
    formulas = sum(1 for n in wbf.sheetnames for row in wbf[n].iter_rows()
                   for c in row if isinstance(c.value, str) and c.value.startswith("="))
    wbf.close()
    res = {"status": "success" if total == 0 else "errors_found",
           "total_errors": total, "total_formulas": formulas, "error_summary": {}}
    for e, locs in details.items():
        if locs:
            res["error_summary"][e] = {"count": len(locs), "locations": locs[:20]}
    return res


def recalc(filename, timeout=30) -> dict:
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}
    soffice = find_soffice()
    if not soffice:
        return {"error": "LibreOffice (soffice) is not installed. Install it and add it to PATH, "
                         "or set SOFFICE_BIN. (Excel calculates the formulas when you open the file.)"}
    if not setup_macro(soffice):
        return {"error": "Could not set up the LibreOffice macro"}
    cmd = [soffice, "--headless", "--norestore",
           "vnd.sun.star.script:Standard.Recalc.RecalcAndSave?language=Basic&location=application",
           str(Path(filename).absolute())]
    try:
        r = subprocess.run(cmd, capture_output=True, encoding="utf-8", errors="replace", timeout=timeout)
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice exceeded the {timeout}s timeout"}
    if r.returncode != 0:
        return {"error": r.stderr or "Unknown error during recalculation"}
    try:
        return scan_errors(filename)
    except Exception as e:  # noqa: BLE001
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("usage: python -X utf8 recalc.py <workbook.xlsx> [timeout_seconds]")
        sys.exit(1)
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc(sys.argv[1], timeout), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
